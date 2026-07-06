"""Web interface for scraping dealers and managing generated S3 files."""

from datetime import datetime
from dataclasses import replace
import importlib
from io import BytesIO
import inspect
import os
from pathlib import Path
import re
from urllib.parse import quote_plus

from dotenv import load_dotenv
import openpyxl
import pandas as pd
import streamlit as st

from catalog import CATEGORY_BRANDS, brands_for_category, canonical_category
from core import (
    PluginRegistry,
    export_to_xlsx,
    google_maps_url,
    pincode_from_address,
    records_with_duplicate_status,
)
from core.s3_storage import S3Storage
from core.indiamart import run_indiamart_scrape
import core.google_places as google_places_module
import core.schema as schema_module

ROOT = Path(__file__).parent.resolve()
OUTPUT_DIR = ROOT / "output"
load_dotenv(ROOT / ".env")


@st.cache_resource
def load_registry() -> PluginRegistry:
    registry = PluginRegistry()
    registry.discover(ROOT / "brands")
    return registry


@st.cache_resource
def load_storage(bucket_name: str, region: str) -> S3Storage:
    return S3Storage(bucket_name=bucket_name, region=region)


def verify_records_with_google_current(
    records,
    *,
    include_unverified: bool,
    **kwargs,
):
    """Verify using a coherent, current schema even after a Streamlit hot reload."""
    kwargs["include_unverified"] = include_unverified
    current_schema = importlib.reload(schema_module)
    current_google_places = importlib.reload(google_places_module)
    field_names = current_schema.DealerRecord.__dataclass_fields__
    current_records = [
        current_schema.DealerRecord(
            **{
                name: getattr(record, name)
                for name in field_names
                if hasattr(record, name)
            }
        )
        for record in records
    ]
    return current_google_places.verify_records_with_google(
        current_records,
        **kwargs,
    )


def verify_records_with_google_including_unverified(records, **kwargs):
    return verify_records_with_google_current(
        records,
        include_unverified=True,
        **kwargs,
    )


def verify_brand_records_with_google_v2_current(records, **kwargs):
    """Run Brand V2 using the current hot-reloaded schema."""
    current_schema = importlib.reload(schema_module)
    current_google_places = importlib.reload(google_places_module)
    field_names = current_schema.DealerRecord.__dataclass_fields__
    current_records = [
        current_schema.DealerRecord(
            **{
                name: getattr(record, name)
                for name in field_names
                if hasattr(record, name)
            }
        )
        for record in records
    ]
    return current_google_places.verify_brand_records_with_google_v2(
        current_records,
        include_unverified=True,
        **kwargs,
    )


def available_brands(registry: PluginRegistry, category: str) -> list[str]:
    installed = {name.casefold(): name for name in registry.list_brands()}
    return [
        installed[name.casefold()]
        for name in brands_for_category(category)
        if name.casefold() in installed
    ]


def active_bucket_name() -> str:
    return str(os.getenv("BUCKET_NAME", "")).strip()


def active_region() -> str:
    return str(os.getenv("AWS_REGION", "us-east-2")).strip()


def safe_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "_", str(value or "").casefold()).strip("_")
    return cleaned or "dealers"


def export_filename(category: str, city: str = "", pincode: str = "") -> str:
    parts = [
        safe_filename_part(city),
        safe_filename_part(pincode),
        safe_filename_part(category),
        datetime.now().strftime("%Y%m%d_%H%M%S"),
    ]
    return "_".join(part for part in parts if part and part != "dealers") + ".xlsx"


def google_terms_for_category(category: str) -> str:
    category_key = str(category or "").casefold()
    if "fan" in category_key:
        return "fans"
    if "cool" in category_key or "roof" in category_key or "paint" in category_key:
        return "paint"
    return "bathroom fitting sanitaryware"


def fetch_records(handler, category: str, state: str, city: str, pincode: str = ""):
    kwargs = {"category": category, "state": state, "city": city}
    if "pincode" in inspect.signature(handler.fetch).parameters:
        kwargs["pincode"] = pincode
    return handler.fetch(**kwargs)


def parse_pincodes(value: str) -> list[str]:
    return list(dict.fromkeys(re.findall(r"(?<!\d)[1-9]\d{5}(?!\d)", value or "")))


def records_for_google_pincodes(records, pincodes, *, keep_other_pincodes: bool):
    """Extract Google-address pincodes, then prioritize/filter requested ones."""
    requested = list(dict.fromkeys(
        str(pin).strip() for pin in pincodes if str(pin).strip()
    ))
    priorities = {pincode: index for index, pincode in enumerate(requested)}
    enriched_records = [
        replace(
            record,
            pincode=(
                pincode_from_address(
                    getattr(record, "google_full_address", "")
                )
                or str(getattr(record, "pincode", "") or "")
            ),
        )
        for record in records
    ]
    if not requested:
        return enriched_records

    def priority(record) -> int:
        return priorities.get(
            str(getattr(record, "pincode", "") or ""),
            len(requested),
        )

    sorted_records = sorted(
        enriched_records,
        key=lambda record: (
            priority(record),
            -float(getattr(record, "google_rating", 0) or 0),
            -int(getattr(record, "google_reviews", 0) or 0),
        ),
    )
    if keep_other_pincodes:
        return sorted_records
    return [
        record for record in sorted_records
        if priority(record) < len(requested)
    ]


BASE_DISPLAY_COLUMNS = {
    "source_brand": "Source Brand",
    "category": "Category",
    "name": "Dealer Name",
    "phone": "Phone",
    "email": "Email",
    "address": "Address",
    "city": "City",
    "state": "State",
    "pincode": "Pincode",
    "dealer_type": "Dealer Type",
    "website": "Website",
    "map_url": "Google Maps",
}

GOOGLE_DISPLAY_COLUMNS = {
    "google_name": "Google Name",
    "google_full_address": "Google Full Address",
    "google_contact_number": "Google Contact Number",
    "google_rating": "Google Rating",
    "google_reviews": "Google Reviews",
    "google_business_type": "Google Business Type",
    "google_business_status": "Google Business Status",
    "google_location": "Google Location",
    "google_name_match_score": "Google Name Match Score",
    "google_distance_km": "Distance from Anchor (km)",
    "google_notes": "Notes",
    "google_pinlocation": "Matched Pinlocation",
    "google_verification_status": "Google Verification Status",
    "google_verification_reason": "Google Verification Reason",
    "google_score": "Google Score",
}

DUPLICATE_STATUS_COLUMN = "Duplicate Status"
HIDDEN_OUTPUT_COLUMNS = {
    "Duplicate Status",
    "State Query",
    "Latitude",
    "Longitude",
    "Google Verified",
    "Google Place ID",
    "Google Get Directions",
    "Google Directions",
}
GOOGLE_OUTPUT_COLUMNS = set(GOOGLE_DISPLAY_COLUMNS.values())
B2B_PRODUCTS = {
    "Flyash Brick": "fly-ash-bricks",
    "AAC Block": "autoclaved-aerated-concrete-block",
    "Cement Brick": "cement-brick",
    "Concrete Block": "concrete-blocks",
    "Hollow Clay Brick": "hollow-clay-bricks",
    "CLC Block": "clc-block",
    "Hollow Concrete Block": "concrete-hollow-blocks",
}

BRAND_V2_DISPLAY_COLUMNS = {
    "source_brand": "Source Brand(s)",
    "name": "Source Dealer Name",
    "phone": "Source Phone",
    "google_full_address": "Matched Address",
    "google_contact_number": "Google Phone",
    "google_rating": "Rating",
    "google_reviews": "Review Count",
    "city": "City",
    "state": "State",
    "pincode": "Pincode",
    "google_distance_km": "Distance from Anchor (km)",
    "google_notes": "Notes",
    "google_pinlocation": "Matched Pinlocation",
}


def is_duplicate_status(value) -> bool:
    return str(value or "").startswith("Duplicate")


def normalized_text(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).strip()


def find_column(dataframe: pd.DataFrame, options: set[str]) -> str | None:
    for column in dataframe.columns:
        if str(column).strip().casefold() in options:
            return column
    return None


def blank_cell(value) -> bool:
    return value is None or pd.isna(value) or not str(value).strip()


def sort_b2b_dataframe_for_pincodes(
    dataframe: pd.DataFrame,
    pincodes: list[str],
) -> pd.DataFrame:
    if dataframe.empty or not pincodes:
        return dataframe
    priorities = {pincode: index for index, pincode in enumerate(pincodes)}

    def priority(row) -> int:
        searchable = " ".join(
            str(row.get(column, "") or "")
            for column in ("Pin Location", "Address", "Google Full Address")
        )
        return min(
            (priorities[pin] for pin in pincodes if pin in searchable),
            default=len(pincodes),
        )

    ordered = dataframe.copy()
    ordered["_pincode_priority"] = dataframe.apply(priority, axis=1)
    ratings = pd.to_numeric(ordered.get("Rating"), errors="coerce").fillna(0)
    reviews = pd.to_numeric(ordered.get("Review Count"), errors="coerce").fillna(0)
    trustseal = ordered.get(
        "Trust Seal / GST Verified",
        pd.Series("", index=ordered.index),
    ).astype(str).str.casefold().eq("yes")
    ordered["_vendor_preferred"] = trustseal | (
        ratings.ge(3.0) & reviews.ge(5)
    )
    ordered["_vendor_rating"] = ratings
    ordered["_vendor_reviews"] = reviews
    ordered = ordered.sort_values(
        [
            "_pincode_priority",
            "_vendor_preferred",
            "_vendor_rating",
            "_vendor_reviews",
            "Company Name",
        ],
        ascending=[True, False, False, False, True],
        kind="stable",
    )
    return ordered.drop(columns=[
        "_pincode_priority",
        "_vendor_preferred",
        "_vendor_rating",
        "_vendor_reviews",
    ]).reset_index(drop=True)


def looks_like_map_url(value) -> bool:
    text = str(value or "").casefold()
    return "google.com/maps" in text or "maps.google" in text or "maps?q=" in text


def enrich_location_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    enriched = dataframe.copy()
    address_column = find_column(enriched, {"address"})
    pincode_column = find_column(enriched, {"pincode", "pin code", "postal code"})

    if address_column:
        if not pincode_column:
            enriched["Pincode"] = ""
            pincode_column = "Pincode"
        enriched[pincode_column] = enriched.apply(
            lambda row: pincode_from_address(row.get(address_column, ""))
            if blank_cell(row.get(pincode_column, ""))
            else row.get(pincode_column, ""),
            axis=1,
        )

    lat_column = find_column(enriched, {"latitude", "lat"})
    lon_column = find_column(enriched, {"longitude", "long", "lng", "lon"})
    maps_column = find_column(enriched, {"google maps", "map url", "map_url", "maps"})
    website_column = find_column(enriched, {"website"})
    if not maps_column:
        enriched["Google Maps"] = ""
        maps_column = "Google Maps"

    if website_column:
        enriched[maps_column] = enriched.apply(
            lambda row: row.get(website_column, "")
            if blank_cell(row.get(maps_column, "")) and looks_like_map_url(row.get(website_column, ""))
            else row.get(maps_column, ""),
            axis=1,
        )
        enriched[website_column] = enriched[website_column].apply(
            lambda value: "" if looks_like_map_url(value) else value
        )

    if lat_column and lon_column:
        if not maps_column:
            enriched["Google Maps"] = ""
            maps_column = "Google Maps"
        enriched[maps_column] = enriched.apply(
            lambda row: google_maps_url(row.get(lat_column, ""), row.get(lon_column, ""))
            if blank_cell(row.get(maps_column, ""))
            else row.get(maps_column, ""),
            axis=1,
        )

    return enriched


def add_duplicate_status_if_possible(dataframe: pd.DataFrame) -> pd.DataFrame:
    if DUPLICATE_STATUS_COLUMN in dataframe.columns:
        return dataframe

    address_column = find_column(dataframe, {"address"})
    if not address_column:
        return dataframe

    keys = []
    counts = {}
    for _, row in dataframe.iterrows():
        address = normalized_text(row.get(address_column, ""))
        key = address if address else None
        keys.append(key)
        if key is not None:
            counts[key] = counts.get(key, 0) + 1

    first_rows = {}
    statuses = []
    for sheet_row, key in enumerate(keys, start=2):
        if key is None or counts.get(key, 0) < 2:
            statuses.append("Not Duplicate")
            continue

        first_row = first_rows.setdefault(key, sheet_row)
        if first_row == sheet_row:
            statuses.append(f"Duplicate (first record, {counts[key]} matches)")
        else:
            statuses.append(f"Duplicate (same as row {first_row})")

    annotated = dataframe.copy()
    annotated.insert(0, DUPLICATE_STATUS_COLUMN, statuses)
    return annotated


def render_duplicate_dataframe(dataframe: pd.DataFrame, preview_limit: int | None = None) -> None:
    dataframe = drop_empty_rows(dataframe)
    dataframe = enrich_location_columns(dataframe)
    dataframe = ensure_google_link_columns(dataframe)
    dataframe = drop_empty_google_columns(dataframe)
    dataframe = add_duplicate_status_if_possible(dataframe)
    duplicate_mask = (
        dataframe[DUPLICATE_STATUS_COLUMN].map(is_duplicate_status)
        if DUPLICATE_STATUS_COLUMN in dataframe.columns
        else pd.Series(False, index=dataframe.index)
    )
    if DUPLICATE_STATUS_COLUMN in dataframe.columns:
        duplicate_count = duplicate_mask.sum()
        not_duplicate_count = len(dataframe) - duplicate_count
        st.caption(
            f"Not duplicate: {not_duplicate_count} | Duplicate rows: {duplicate_count}"
        )

    def mark_duplicates(row):
        if bool(duplicate_mask.get(row.name, False)):
            return ["background-color: #f4cccc; color: #990000"] * len(row)
        return [""] * len(row)

    if preview_limit is not None:
        st.caption(f"Previewing up to {preview_limit} rows")

    visible_dataframe = dataframe.drop(
        columns=[column for column in HIDDEN_OUTPUT_COLUMNS if column in dataframe.columns]
    )
    column_config = {}
    if "Google Maps" in visible_dataframe.columns:
        column_config["Google Maps"] = st.column_config.LinkColumn(
            "Google Maps",
        )
    if "Google Location" in visible_dataframe.columns:
        column_config["Google Location"] = st.column_config.LinkColumn(
            "Google Location",
            display_text="Open Map",
        )
    if "Matched Pinlocation" in visible_dataframe.columns:
        column_config["Matched Pinlocation"] = st.column_config.LinkColumn(
            "Matched Pinlocation",
        )
    if "Website" in visible_dataframe.columns:
        column_config["Website"] = st.column_config.LinkColumn("Website")

    link_columns = {
        "Google Maps",
        "Google Location",
        "Matched Pinlocation",
        "Website",
    }
    if link_columns & set(visible_dataframe.columns):
        table = visible_dataframe
    else:
        table = visible_dataframe.style.apply(mark_duplicates, axis=1)

    st.dataframe(
        table,
        use_container_width=True,
        hide_index=True,
        column_config=column_config or None,
    )


def dataframe_without_duplicate_rows(dataframe: pd.DataFrame) -> pd.DataFrame:
    annotated = add_duplicate_status_if_possible(
        drop_empty_rows(drop_empty_google_columns(dataframe))
    )
    if DUPLICATE_STATUS_COLUMN not in annotated.columns:
        return dataframe.drop(
            columns=[column for column in HIDDEN_OUTPUT_COLUMNS if column in dataframe.columns]
        )
    duplicate_statuses = annotated[DUPLICATE_STATUS_COLUMN].astype(str)
    keep_rows = ~duplicate_statuses.str.startswith("Duplicate (same as row")
    deduped = annotated.loc[keep_rows]
    return deduped.drop(
        columns=[column for column in HIDDEN_OUTPUT_COLUMNS if column in deduped.columns]
    )


def dataframe_without_unverified_rows(dataframe: pd.DataFrame) -> pd.DataFrame:
    status_column = find_column(dataframe, {"google verification status"})
    if not status_column:
        return dataframe
    statuses = dataframe[status_column].astype(str).str.strip().str.casefold()
    return dataframe.loc[statuses.ne("unverified")]


def dataframe_with_download_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    enriched = enrich_location_columns(dataframe)
    enriched = ensure_google_link_columns(enriched)
    enriched = drop_empty_google_columns(enriched)
    enriched = drop_empty_rows(enriched)
    return enriched.drop(
        columns=[column for column in HIDDEN_OUTPUT_COLUMNS if column in enriched.columns]
    )


def drop_empty_rows(dataframe: pd.DataFrame) -> pd.DataFrame:
    if dataframe.empty:
        return dataframe
    cleaned = dataframe.replace(r"^\s*$", pd.NA, regex=True)
    return cleaned.dropna(how="all").fillna("")


def ensure_google_link_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    enriched = dataframe.copy()
    address_column = find_column(enriched, {"google full address"})
    name_column = find_column(enriched, {"google name"})
    location_column = find_column(enriched, {"google location"})

    if not address_column and not name_column:
        return enriched

    def destination(row) -> str:
        address = row.get(address_column, "") if address_column else ""
        name = row.get(name_column, "") if name_column else ""
        return str(address or name or "").strip()

    if not location_column:
        enriched["Google Location"] = ""
        location_column = "Google Location"
    enriched[location_column] = enriched.apply(
        lambda row: (
            f"https://www.google.com/maps/search/?api=1&query={quote_plus(destination(row))}"
            if blank_cell(row.get(location_column, "")) and destination(row)
            else row.get(location_column, "")
        ),
        axis=1,
    )

    return enriched


def dataframe_to_xlsx_bytes(dataframe: pd.DataFrame) -> bytes:
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Dealer Logbook"
    for col_index, column in enumerate(dataframe.columns, start=1):
        sheet.cell(row=1, column=col_index, value=str(column))
    for row_index, (_, row) in enumerate(dataframe.iterrows(), start=2):
        for col_index, (column, value) in enumerate(row.items(), start=1):
            cell_value = "" if pd.isna(value) else value
            cell = sheet.cell(row=row_index, column=col_index, value=cell_value)
            if (
                str(column) in {"Google Maps", "Matched Pinlocation"}
                and looks_like_map_url(cell_value)
            ):
                cell.hyperlink = str(cell_value)
                cell.style = "Hyperlink"
    workbook.save(output)
    return output.getvalue()


def saved_file_download_payload(
    key: str,
    data: bytes,
    *,
    without_duplicates: bool,
    include_unverified: bool,
) -> tuple[bytes, str, str]:
    suffix = Path(key).suffix.casefold()
    name = Path(key).name
    stem = Path(key).stem
    if suffix == ".xlsx":
        workbook = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        if not rows:
            return data, name, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        headers = [str(value or "") for value in rows[0]]
        dataframe = pd.DataFrame([dict(zip(headers, row)) for row in rows[1:]])
        clean_dataframe = dataframe_with_download_columns(dataframe)
        if without_duplicates:
            clean_dataframe = dataframe_without_duplicate_rows(clean_dataframe)
        if not include_unverified:
            clean_dataframe = dataframe_without_unverified_rows(clean_dataframe)
        suffix_parts = []
        if without_duplicates:
            suffix_parts.append("without_duplicates")
        if not include_unverified:
            suffix_parts.append("verified_only")
        download_name = f"{stem}_{'_'.join(suffix_parts)}.xlsx" if suffix_parts else name
        return (
            dataframe_to_xlsx_bytes(clean_dataframe),
            download_name,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    if suffix == ".csv":
        dataframe = pd.read_csv(BytesIO(data))
        clean_dataframe = dataframe_with_download_columns(dataframe)
        if without_duplicates:
            clean_dataframe = dataframe_without_duplicate_rows(clean_dataframe)
        if not include_unverified:
            clean_dataframe = dataframe_without_unverified_rows(clean_dataframe)
        suffix_parts = []
        if without_duplicates:
            suffix_parts.append("without_duplicates")
        if not include_unverified:
            suffix_parts.append("verified_only")
        download_name = f"{stem}_{'_'.join(suffix_parts)}.csv" if suffix_parts else name
        return (
            clean_dataframe.to_csv(index=False).encode("utf-8"),
            download_name,
            "text/csv",
        )
    return data, name, "application/octet-stream"


def drop_empty_google_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    google_columns = [
        column for column in GOOGLE_OUTPUT_COLUMNS
        if column in dataframe.columns
        and dataframe[column].map(lambda value: not blank_cell(value)).sum() == 0
    ]
    if not google_columns:
        return dataframe
    return dataframe.drop(columns=google_columns)


def display_columns_for_records(records) -> dict:
    has_google_data = any(record.google_verified is not None for record in records)
    return (
        {**BASE_DISPLAY_COLUMNS, **GOOGLE_DISPLAY_COLUMNS}
        if has_google_data
        else BASE_DISPLAY_COLUMNS
    )


def render_records_table(records) -> None:
    rows = records_with_duplicate_status(records)
    dataframe = pd.DataFrame(rows)
    display_columns = display_columns_for_records(records)
    dataframe = dataframe[[column for column in display_columns if column in dataframe.columns]]
    dataframe = dataframe.rename(columns=display_columns)
    render_duplicate_dataframe(dataframe)


def render_brand_v2_table(records) -> None:
    dataframe = pd.DataFrame([record.to_dict() for record in records])
    columns = [
        column for column in BRAND_V2_DISPLAY_COLUMNS
        if column in dataframe.columns
    ]
    dataframe = dataframe[columns].rename(columns=BRAND_V2_DISPLAY_COLUMNS)
    render_duplicate_dataframe(dataframe)


def records_download_dataframe(records) -> pd.DataFrame:
    rows = records_with_duplicate_status(records)
    dataframe = pd.DataFrame(rows)
    display_columns = display_columns_for_records(records)
    dataframe = dataframe[[column for column in display_columns if column in dataframe.columns]]
    dataframe = dataframe.rename(columns=display_columns)
    return enrich_location_columns(dataframe)


def brand_v2_download_dataframe(records) -> pd.DataFrame:
    """Build the compact post-verification brand download."""
    columns = [
        "Category",
        "Source Dealer Name",
        "Source Phone",
        "Google Phone",
        "Matched Address",
        "City",
        "Matched Pinlocation",
    ]
    verified = [
        record for record in records
        if bool(getattr(record, "google_verified", False))
    ]
    return pd.DataFrame([
        {
            "Category": record.category,
            "Source Dealer Name": record.name,
            "Source Phone": record.phone,
            "Google Phone": record.google_contact_number,
            "Matched Address": record.google_full_address,
            "City": record.city,
            "Matched Pinlocation": record.google_pinlocation,
        }
        for record in verified
    ], columns=columns)


def save_records_to_shared_files(
    records,
    *,
    filename: str,
    bucket_name: str,
) -> tuple[Path, str | None, str | None]:
    output_path = export_to_xlsx(records, OUTPUT_DIR, filename)
    if not bucket_name:
        return output_path, None, None
    try:
        storage = load_storage(bucket_name, active_region())
        key = storage.upload_path(output_path, key=f"exports/{filename}")
        return output_path, key, None
    except Exception as exc:
        return output_path, None, str(exc)


def save_brand_v2_to_shared_files(
    records,
    *,
    filename: str,
    bucket_name: str,
) -> tuple[Path, str | None, str | None]:
    output_path = OUTPUT_DIR / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(
        dataframe_to_xlsx_bytes(brand_v2_download_dataframe(records))
    )
    if not bucket_name:
        return output_path, None, None
    try:
        storage = load_storage(bucket_name, active_region())
        key = storage.upload_path(output_path, key=f"exports/{filename}")
        return output_path, key, None
    except Exception as exc:
        return output_path, None, str(exc)


def render_scraper_dashboard(registry: PluginRegistry) -> None:
    category = st.selectbox("Category", list(CATEGORY_BRANDS))
    scraper_category = canonical_category(category)
    brands = available_brands(registry, category)
    configured = brands_for_category(category)
    missing = [name for name in configured if registry.get(name) is None]

    if not brands:
        st.warning("No sources are available for this category yet.")
        if missing:
            st.caption("Configured, but not implemented: " + ", ".join(missing))
        return

    selected_brands = st.multiselect("Brands", brands, default=brands)
    if missing:
        st.caption("Coming later (no plugin installed): " + ", ".join(missing))

    left, middle, right = st.columns(3)
    state = left.text_input("State", placeholder="Karnataka")
    city = middle.text_input("City", placeholder="Bengaluru")
    pincode_text = right.text_input(
        "Pincodes *",
        placeholder="560001, 560002",
        help="Enter one or more six-digit pincodes separated by commas or spaces.",
    )
    pincodes = parse_pincodes(pincode_text)
    city_required = [
        brand for brand in selected_brands
        if registry.get(brand) and registry.get(brand).REQUIRES_CITY
    ]
    if city_required:
        st.caption("City is required for: " + ", ".join(city_required))

    bucket_name = active_bucket_name()
    upload_to_s3 = bool(bucket_name)

    scrape_signature = {
        "category": category,
        "scraper_category": scraper_category,
        "brands": tuple(selected_brands),
        "state": state.strip(),
        "city": city.strip(),
        "pincodes": tuple(pincodes),
    }

    if st.button(
        "Run Brand Authorized Directory",
        type="primary",
        use_container_width=True,
    ):
        errors: list[str] = []
        records = []
        if not selected_brands:
            st.error("Select at least one brand.")
            return
        if not state.strip():
            st.error("State is required.")
            return
        if city_required and not city.strip():
            st.error("City is required for: " + ", ".join(city_required))
            return
        invalid_pincode_text = re.sub(r"[\s,;]+", "", pincode_text)
        if not pincodes or invalid_pincode_text != "".join(pincodes):
            st.error("Enter valid six-digit pincodes separated by commas or spaces.")
            return

        jobs = []
        for brand in selected_brands:
            handler_class = registry.get(brand)
            brand_pincodes = (
                pincodes
                if handler_class and getattr(handler_class, "REQUIRES_PINCODE", False)
                else [pincodes[0]]
            )
            jobs.extend((brand, pin) for pin in brand_pincodes)

        progress = st.progress(0, text="Starting...")
        for index, (brand, active_pincode) in enumerate(jobs, start=1):
            progress.progress(
                (index - 1) / len(jobs),
                text=f"Running {brand} for {active_pincode}...",
            )
            try:
                handler_class = registry.get(brand)
                if handler_class is None:
                    raise LookupError(f"No source is installed for {brand}.")
                records.extend(fetch_records(
                    handler_class(),
                    category=scraper_category,
                    state=state.strip(),
                    city=city.strip(),
                    pincode=active_pincode,
                ))
            except Exception as exc:
                errors.append(f"{brand} ({active_pincode}): {exc}")
            progress.progress(index / len(jobs), text=f"Finished {brand}")
        progress.empty()
        raw_count = len(records)
        filename = export_filename(category, city.strip(), "-".join(pincodes))
        _, saved_key, save_error = save_records_to_shared_files(
            records,
            filename=filename,
            bucket_name=bucket_name,
        )

        st.session_state.pop("google_verified_result", None)
        st.session_state["scraper_result"] = {
            "signature": scrape_signature,
            "records": records,
            "errors": errors,
            "filename": filename,
            "raw_count": raw_count,
            "saved_key": saved_key,
            "save_error": save_error,
        }

    result = st.session_state.get("scraper_result")
    if not result or result.get("signature") != scrape_signature:
        return

    all_records = result.get("records", [])
    errors = result.get("errors", [])
    if errors:
        st.warning("Some sources could not finish:\n\n" + "\n\n".join(
            f"- {error}" for error in errors
        ))
    if not all_records:
        if not errors:
            st.info("No dealer records were returned.")
        return

    keep_other_pincodes = st.checkbox(
        "After Google verification, keep dealers from other pincodes",
        value=False,
        key="keep-other-pincodes-after-google",
        help=(
            "Google Full Address supplies the pincode. When enabled, unmatched "
            "or other-pincode dealers remain after requested-pincode results."
        ),
    )
    records = all_records

    st.success(f"Found {len(records)} dealer records across {len(selected_brands)} brand(s).")
    if result.get("raw_count", len(all_records)) != len(all_records):
        st.caption(
            f"Removed {result.get('raw_count', len(all_records)) - len(all_records)} duplicate address row(s)."
        )
    if result.get("saved_key"):
        st.caption(f"Saved to shared files: {Path(result['saved_key']).name}")
    elif result.get("save_error"):
        st.warning(f"Local Excel was generated, but shared file save failed: {result['save_error']}")
    render_records_table(records)

    st.divider()
    if st.button(
        "Verify with Google Places",
        type="primary",
        use_container_width=True,
        help=(
            "Uses coordinates, pincode, or road/city geocoding as an anchor; "
            "then independently checks distance, address overlap, and name."
        ),
    ):
        before_count = len(records)
        verify_progress = st.progress(0, text="Starting Google Places verification...")

        def update_google_progress(index: int, total: int, record) -> None:
            name = str(getattr(record, "name", "") or "dealer")
            verify_progress.progress(index / total, text=f"Checking {name}...")

        try:
            verified_records = verify_brand_records_with_google_v2_current(
                records,
                state=state.strip(),
                city=city.strip(),
                on_progress=update_google_progress,
            )
        except Exception as exc:
            verify_progress.empty()
            st.error(f"Google Places verification failed: {exc}")
            return
        verify_progress.empty()
        verified_records = records_for_google_pincodes(
            verified_records,
            pincodes,
            keep_other_pincodes=keep_other_pincodes,
        )
        _, verified_saved_key, verified_save_error = save_brand_v2_to_shared_files(
            verified_records,
            filename=result["filename"],
            bucket_name=bucket_name,
        )
        st.session_state["google_verified_result"] = {
            "signature": scrape_signature,
            "records": verified_records,
            "source_count": before_count,
            "filename": result["filename"],
            "saved_key": verified_saved_key,
            "save_error": verified_save_error,
        }

    verified_result = st.session_state.get("google_verified_result")
    if not verified_result or verified_result.get("signature") != scrape_signature:
        return

    verified_records = verified_result.get("records", [])
    verified_count = sum(
        1 for record in verified_records
        if str(getattr(record, "google_verification_status", "") or "").casefold() == "verified"
    )
    unverified_count = len(verified_records) - verified_count
    st.info(
        f"Google Places verified {verified_count} of "
        f"{verified_result.get('source_count', len(records))} records. "
        f"Unverified rows shown: {unverified_count}."
    )
    if not verified_records:
        if keep_other_pincodes:
            st.warning("Google Places did not return any dealer records.")
        else:
            st.warning(
                "No Google Full Address matched the requested pincode(s): "
                + ", ".join(pincodes)
            )
        return

    render_brand_v2_table(verified_records)
    if verified_result.get("saved_key"):
        st.caption(f"Replaced saved file: {Path(verified_result['saved_key']).name}")
    elif verified_result.get("save_error"):
        st.warning(f"Verified Excel was generated, but shared file save failed: {verified_result['save_error']}")


def render_s3_dashboard() -> None:
    st.subheader("Output")
    region = active_region()
    bucket_name = active_bucket_name()
    if not bucket_name:
        st.info("Shared file storage is not configured yet.")
        return

    try:
        storage = load_storage(bucket_name, region)
        exists = storage.bucket_exists()
    except Exception as exc:
        st.error(f"Could not connect to shared file storage: {exc}")
        return

    if not exists:
        st.warning("Shared file storage is not available.")
        return

    try:
        files = storage.list_files()
    except Exception as exc:
        st.error(f"Could not list files: {exc}")
        return
    if not files:
        st.info("The exports/ folder is empty.")
        return

    st.dataframe([
        {
            "File": item.name,
            "Size (KB)": round(item.size / 1024, 2),
            "Last modified": item.last_modified,
        }
        for item in files
    ], use_container_width=True, hide_index=True)

    file_options = {item.name: item.key for item in files}
    selected_name = st.selectbox("Select a file", list(file_options))
    selected_key = file_options[selected_name]
    if st.button("Retrieve and preview"):
        try:
            st.session_state["s3_selected_data"] = storage.download_bytes(selected_key)
            st.session_state["s3_selected_key"] = selected_key
        except Exception as exc:
            st.error(f"Retrieval failed: {exc}")

    if st.session_state.get("s3_selected_key") == selected_key:
        data = st.session_state.get("s3_selected_data", b"")
        include_duplicates = st.checkbox(
            "Include duplicate rows",
            value=True,
            key=f"include-duplicates-{selected_key}",
        )
        include_unverified = st.checkbox(
            "Include unverified Google rows",
            value=True,
            key=f"include-unverified-{selected_key}",
        )
        try:
            download_data, download_name, download_mime = saved_file_download_payload(
                selected_key,
                data,
                without_duplicates=not include_duplicates,
                include_unverified=include_unverified,
            )
        except Exception as exc:
            st.warning(f"Could not prepare download: {exc}")
            download_data = data
            download_name = Path(selected_key).name
            download_mime = "application/octet-stream"
        st.download_button(
            "Download file",
            data=download_data,
            file_name=download_name,
            mime=download_mime,
            key=f"download-selected-{selected_key}",
            use_container_width=True,
        )
        st.caption("The preview below reflects the selected row filters.")
        render_file_preview(download_name, download_data)

    confirm_delete = st.checkbox("I confirm deletion of the selected file")
    if st.button("Delete selected file", disabled=not confirm_delete):
        try:
            storage.delete_file(selected_key)
            st.session_state.pop("s3_selected_data", None)
            st.session_state.pop("s3_selected_key", None)
            st.success(f"Deleted {selected_key}")
            st.rerun()
        except Exception as exc:
            st.error(f"Deletion failed: {exc}")


def render_indiamart_dashboard() -> None:
    st.subheader("B2B Directory")
    st.caption(
        "Scrape IndiaMART listings in a visible browser (debug mode). "
        "Listings are retained when a TrustSEAL is found on the card or company "
        "page, or when they have at least 3 stars and 5 reviews."
    )

    left, right = st.columns(2)
    city = left.text_input(
        "City",
        value="Indore",
    )
    selected_products = right.multiselect(
        "Products",
        options=list(B2B_PRODUCTS),
        default=["Flyash Brick"],
    )
    b2b_pincode_text = st.text_input(
        "Preferred pincodes for sorting",
        placeholder="452001, 452002",
        help="These pincodes sort B2B results only; they do not change the IndiaMART search.",
    )
    b2b_pincodes = parse_pincodes(b2b_pincode_text)
    st.caption("Porotherm and the remaining planned products will be added when their slugs are available.")

    if st.button(
        "Run B2B Directory",
        type="primary",
        use_container_width=True,
    ):
        progress = st.progress(0, text="Starting IndiaMART scraper...")

        def update_progress(done: int, total: int, label: str) -> None:
            progress.progress(
                done / total if total else 0,
                text=f"IndiaMART: {label}",
            )

        try:
            if not city.strip():
                raise ValueError("Enter a city.")
            if not selected_products:
                raise ValueError("Select at least one product.")
            rows = run_indiamart_scrape(
                [city.strip()],
                [B2B_PRODUCTS[product] for product in selected_products],
                apply_quality_gate=True,
                on_progress=update_progress,
            )
        except Exception as exc:
            progress.empty()
            st.error(f"IndiaMART scraping failed: {exc}")
        else:
            progress.empty()
            dataframe = sort_b2b_dataframe_for_pincodes(
                pd.DataFrame(rows),
                b2b_pincodes,
            )
            csv_data = dataframe.to_csv(index=False).encode("utf-8-sig")
            filename = (
                "indiamart_leads_"
                + datetime.now().strftime("%Y%m%d_%H%M%S")
                + ".csv"
            )
            output_path = OUTPUT_DIR / filename
            output_path.write_bytes(csv_data)
            saved_key = None
            save_error = None
            if active_bucket_name():
                try:
                    storage = load_storage(active_bucket_name(), active_region())
                    saved_key = storage.upload_path(
                        output_path,
                        key=f"exports/{filename}",
                    )
                except Exception as exc:
                    save_error = str(exc)
            st.session_state["indiamart_result"] = {
                "dataframe": dataframe,
                "csv_data": csv_data,
                "filename": filename,
                "saved_key": saved_key,
                "save_error": save_error,
            }

    result = st.session_state.get("indiamart_result")
    if not result:
        return
    dataframe = result["dataframe"].drop(
        columns=["Product URL", "Product", "Company URL", "Website"],
        errors="ignore",
    )
    dataframe = sort_b2b_dataframe_for_pincodes(dataframe, b2b_pincodes)
    if len(dataframe.columns) != len(result["dataframe"].columns):
        result["dataframe"] = dataframe
        result["csv_data"] = dataframe.to_csv(index=False).encode("utf-8-sig")
    status_column = (
        dataframe["Google Verification Status"].astype(str).str.casefold()
        if "Google Verification Status" in dataframe.columns
        else pd.Series("", index=dataframe.index)
    )
    verified_count = int(status_column.eq("verified").sum())
    unverified_count = len(dataframe) - verified_count
    st.success(f"Found {len(dataframe)} IndiaMART leads.")
    verified_metric, unverified_metric = st.columns(2)
    verified_metric.metric("Verified vendors", verified_count)
    unverified_metric.metric("Unverified vendors", unverified_count)
    if result.get("saved_key"):
        st.caption(f"Saved to shared files: {Path(result['saved_key']).name}")
    elif result.get("save_error"):
        st.warning(f"Local CSV created, but shared save failed: {result['save_error']}")

    google_search_terms = st.text_input(
        "Google Places search terms",
        value="building materials",
        key="indiamart-google-search-terms",
    )
    if st.button(
        "Verify",
        type="primary",
        use_container_width=True,
        disabled=dataframe.empty,
    ):
        source_rows = dataframe.to_dict(orient="records")
        dealer_records = [
            schema_module.DealerRecord(
                name=str(row.get("Company Name", "") or ""),
                address=str(row.get("Address", "") or ""),
                city=str(row.get("City", "") or ""),
                category=str(row.get("Subcategory", "") or ""),
                phone=str(row.get("Phone", "") or ""),
                website=str(row.get("Company URL", "") or ""),
            )
            for row in source_rows
        ]
        verify_progress = st.progress(
            0,
            text="Starting Google Places verification...",
        )

        def update_google_progress(index: int, total: int, record) -> None:
            verify_progress.progress(
                index / total if total else 0,
                text=f"Checking {getattr(record, 'name', '') or 'IndiaMART lead'}...",
            )

        try:
            verified_records = verify_records_with_google_current(
                dealer_records,
                include_unverified=True,
                search_terms=google_search_terms.strip() or "building materials",
                on_progress=update_google_progress,
            )
        except Exception as exc:
            verify_progress.empty()
            st.error(f"Google Places verification failed: {exc}")
        else:
            verify_progress.empty()
            source_by_key = {
                (
                    str(row.get("Company Name", "") or "").strip().casefold(),
                    str(row.get("Address", "") or "").strip().casefold(),
                ): row
                for row in source_rows
            }
            verified_rows = []
            for record in verified_records:
                key = (
                    str(record.name or "").strip().casefold(),
                    str(record.address or "").strip().casefold(),
                )
                row = dict(source_by_key.get(key, {}))
                record_data = record.to_dict()
                for field, label in GOOGLE_DISPLAY_COLUMNS.items():
                    row[label] = record_data.get(field)
                verified_rows.append(row)

            verified_dataframe = pd.DataFrame(verified_rows)
            verified_csv = verified_dataframe.to_csv(index=False).encode("utf-8-sig")
            verified_filename = (
                Path(result["filename"]).stem + "_google_checked.csv"
            )
            verified_path = OUTPUT_DIR / verified_filename
            verified_path.write_bytes(verified_csv)
            saved_key = None
            save_error = None
            if active_bucket_name():
                try:
                    storage = load_storage(active_bucket_name(), active_region())
                    saved_key = storage.upload_path(
                        verified_path,
                        key=f"exports/{verified_filename}",
                    )
                except Exception as exc:
                    save_error = str(exc)
            st.session_state["indiamart_result"] = {
                "dataframe": verified_dataframe,
                "csv_data": verified_csv,
                "filename": verified_filename,
                "saved_key": saved_key,
                "save_error": save_error,
                "google_verified": True,
            }
            st.rerun()

    if result.get("google_verified"):
        st.info(
            "This result includes both verified and unverified Google Places rows."
        )
    st.dataframe(dataframe, use_container_width=True, hide_index=True)


def render_file_preview(key: str, data: bytes) -> None:
    suffix = Path(key).suffix.casefold()
    if suffix == ".xlsx":
        try:
            workbook = openpyxl.load_workbook(BytesIO(data), read_only=False, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(max_row=201))
            if rows:
                headers = [str(cell.value or "") for cell in rows[0]]
                preview_rows = []
                for row in rows[1:]:
                    values = [
                        cell.hyperlink.target
                        if cell.hyperlink and cell.hyperlink.target
                        else cell.value
                        for cell in row
                    ]
                    preview_rows.append(dict(zip(headers, values)))
                dataframe = pd.DataFrame(preview_rows)
                render_duplicate_dataframe(dataframe, preview_limit=200)
        except Exception as exc:
            st.warning(f"Excel preview unavailable: {exc}")
    elif suffix == ".csv":
        try:
            dataframe = pd.read_csv(BytesIO(data))
            render_duplicate_dataframe(dataframe, preview_limit=200)
        except Exception as exc:
            st.warning(f"CSV preview unavailable: {exc}")
    elif suffix in {".json", ".txt"}:
        st.text_area("Preview", data[:100_000].decode("utf-8", errors="replace"), height=300)
    else:
        st.caption("Preview is unavailable for this file type; use Download instead.")


st.set_page_config(page_title="Dealer Finder", page_icon="🔎", layout="wide")
st.markdown(
    """
    <style>
    .stButton button[kind="primary"] {
        background-color: #16803c;
        border-color: #16803c;
        color: #ffffff;
    }
    .stButton button[kind="primary"]:hover {
        background-color: #0f6b31;
        border-color: #0f6b31;
        color: #ffffff;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
st.title("Dealer Finder")
st.caption("Run dealer and IndiaMART scrapers and manage generated files.")

scraper_tab, indiamart_tab, files_tab = st.tabs(
    ["Brand Authorized Directory", "B2B Directory", "Output"]
)
with scraper_tab:
    render_scraper_dashboard(load_registry())
with indiamart_tab:
    render_indiamart_dashboard()
with files_tab:
    render_s3_dashboard()
