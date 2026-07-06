"""
Exports a list of DealerRecord objects to an xlsx logbook.
"""
from datetime import datetime
from pathlib import Path
import re
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .schema import DealerRecord

COLUMNS = [
    ("Source Brand",    "source_brand"),
    ("Category",        "category"),
    ("Dealer Name",     "name"),
    ("Phone",           "phone"),
    ("Email",           "email"),
    ("Address",         "address"),
    ("City",            "city"),
    ("State",           "state"),
    ("Pincode",         "pincode"),
    ("Dealer Type",     "dealer_type"),
    ("Website",         "website"),
    ("Google Maps",     "map_url"),
]

GOOGLE_COLUMNS = [
    ("Google Name",     "google_name"),
    ("Google Full Address", "google_full_address"),
    ("Google Contact Number", "google_contact_number"),
    ("Google Rating",   "google_rating"),
    ("Google Reviews",  "google_reviews"),
    ("Google Business Type", "google_business_type"),
    ("Google Business Status", "google_business_status"),
    ("Google Location", "google_location"),
    ("Google Name Match Score", "google_name_match_score"),
    ("Distance from Anchor (km)", "google_distance_km"),
    ("Notes", "google_notes"),
    ("Matched Pinlocation", "google_pinlocation"),
    ("Google Verification Status", "google_verification_status"),
    ("Google Verification Reason", "google_verification_reason"),
    ("Google Score",    "google_score"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
DUPLICATE_FILL = PatternFill("solid", fgColor="F4CCCC")
DUPLICATE_FONT = Font(color="990000")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
PINCODE_RE = re.compile(r"(?<!\d)([1-9]\d{5})(?!\d)")


def _normalize_duplicate_value(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).strip()


def _duplicate_key(record: DealerRecord) -> tuple[str, str, str] | None:
    name = _normalize_duplicate_value(record.name)
    address = _normalize_duplicate_value(record.address)
    if not name or not address:
        return None
    return ("name_address", name, address)


def _duplicate_statuses(records: List[DealerRecord]) -> list[str]:
    first_rows: dict[tuple[str, str, str], int] = {}
    counts: dict[tuple[str, str, str], int] = {}
    keys = [_duplicate_key(record) for record in records]

    for key in keys:
        if key is not None:
            counts[key] = counts.get(key, 0) + 1

    statuses = []
    for index, key in enumerate(keys, start=2):
        if key is None or counts.get(key, 0) < 2:
            statuses.append("Not Duplicate")
            continue

        first_row = first_rows.setdefault(key, index)
        if first_row == index:
            statuses.append(f"Duplicate (first record, {counts[key]} matches)")
        else:
            statuses.append(f"Duplicate (same as row {first_row})")

    return statuses


def pincode_from_address(address: str) -> str:
    match = PINCODE_RE.search(str(address or ""))
    return match.group(1) if match else ""


def google_maps_url(latitude, longitude) -> str:
    lat = str(latitude or "").strip()
    lon = str(longitude or "").strip()
    if not lat or not lon:
        return ""
    try:
        float(lat)
        float(lon)
    except ValueError:
        return ""
    return f"https://www.google.com/maps?q={lat},{lon}"


def looks_like_map_url(value: str) -> bool:
    url = str(value or "").casefold()
    return "google.com/maps" in url or "maps.google" in url or "maps?q=" in url


def looks_like_locator_url(value: str) -> bool:
    url = str(value or "").casefold()
    return "locator" in url or "find-dealer" in url or "dealer-locator" in url


def export_row_dict(record: DealerRecord) -> dict:
    row = record.to_dict()
    if not str(row.get("pincode") or "").strip():
        row["pincode"] = pincode_from_address(row.get("address", ""))
    website = str(row.get("website") or "").strip()
    map_url = str(row.get("map_url") or "").strip()
    if not map_url and looks_like_map_url(website):
        map_url = website
    if looks_like_map_url(website) or looks_like_locator_url(website):
        row["website"] = ""
    row["map_url"] = map_url or google_maps_url(row.get("latitude"), row.get("longitude"))
    return row


def records_with_duplicate_status(records: List[DealerRecord]) -> list[dict]:
    statuses = _duplicate_statuses(records)
    rows = []
    for record, status in zip(records, statuses):
        row = export_row_dict(record)
        row["duplicate_status"] = status
        rows.append(row)
    return rows


def records_without_duplicates(records: List[DealerRecord]) -> list[DealerRecord]:
    seen: set[tuple[str, str, str]] = set()
    unique_records = []
    for record in records:
        key = _duplicate_key(record)
        if key is None:
            unique_records.append(record)
            continue
        if key in seen:
            continue
        seen.add(key)
        unique_records.append(record)
    return unique_records


def export_columns(records: List[DealerRecord]) -> list[tuple[str, str]]:
    has_google_data = any(record.google_verified is not None for record in records)
    return COLUMNS + GOOGLE_COLUMNS if has_google_data else COLUMNS


def export_to_xlsx(records: List[DealerRecord], output_dir: Path, filename: str = "") -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dealers_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dealer Logbook"
    columns = export_columns(records)

    # ── Header row ────────────────────────────────────────────────────
    for col_idx, (header, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────
    for row_idx, rec_dict in enumerate(records_with_duplicate_status(records), start=2):
        duplicate_status = rec_dict["duplicate_status"]
        is_duplicate = duplicate_status.startswith("Duplicate")
        fill = DUPLICATE_FILL if is_duplicate else ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, (_, field_key) in enumerate(columns, start=1):
            val = rec_dict.get(field_key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val or "")
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if field_key == "map_url" and val:
                cell.hyperlink = val
                cell.value = val
                cell.font = Font(color="0563C1", underline="single")
            if field_key == "google_location" and val:
                cell.hyperlink = val
                cell.value = "Open Map"
                cell.font = Font(color="0563C1", underline="single")
            if fill:
                cell.fill = fill
            if is_duplicate:
                cell.font = DUPLICATE_FONT

    # ── Column widths ─────────────────────────────────────────────────
    col_widths = [
        16, 22, 28, 18, 28, 38, 18, 18, 12, 20, 28, 18,
        28, 42, 22, 14, 15, 30, 22, 28, 24, 36, 14,
    ]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes & auto-filter ────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out_path = output_dir / filename
    wb.save(out_path)
    print(f"[Exporter] Saved {len(records)} records -> {out_path}")
    return out_path
